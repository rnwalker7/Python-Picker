import random
import sys
import glob
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

uua = []
uub = []

a = open("uua.txt", "r")

for x in a:
    uua.append(x)
a.close()

b = open("uub.txt", "r")

for x in b:
    uub.append(x)
b.close()

random.shuffle(uua)
random.shuffle(uub)

# Now for the Excel fun!

dest_filename = 'NCOD Schedule.xlsx'

wb = load_workbook(filename = dest_filename)

ws1 = wb.active

for i in range(len(uua)):
    uua_sub = "B" + str(i + 3)
    uub_sub = "C" + str(i + 3)
    ws1[uua_sub] = uua[i]
    ws1[uub_sub] = uub[i]

wb.save(filename = dest_filename)